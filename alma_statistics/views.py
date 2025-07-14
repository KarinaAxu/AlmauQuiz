from django.shortcuts import render, redirect
from .models import Applicant, QuizResult
from datetime import datetime
import json
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import openpyxl
from pathlib import Path
from django.conf import settings
from rest_framework import generics
from .models import QuizQuestion
from rest_framework.response import Response
from rest_framework.views import APIView

class QuizAPIView(APIView):
    def get(self, request):
        data = []
        for question in QuizQuestion.objects.prefetch_related('answers').all():
            data.append({
                'question': question.question_text,
                'answers': [{'icon': a.icon, 'text': a.text} for a in question.answers.all()]
            })
        return Response(data)

# Путь до media/results.xlsx
MEDIA_DIR = settings.BASE_DIR / 'media'
EXCEL_PATH = MEDIA_DIR / 'results.xlsx'


def init_excel():
    # Создание папки media если нет
    MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    print(f"===> Проверка Excel: {EXCEL_PATH}")

    if not EXCEL_PATH.exists():
        print("===> Создание файла results.xlsx")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Анкеты'
        ws1.append(['ФИО', 'Телефон', 'Школа', 'Источник', 'Номер заявки', 'Дата регистрации', 'Время'])

        ws2 = wb.create_sheet('Викторины')
        ws2.append(['ФИО', 'Ответы', 'Дата и время'])

        wb.save(EXCEL_PATH)
        print(f"===> Файл создан: {EXCEL_PATH}")


@csrf_exempt
def save_full_data(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(f"===> Получены данные: {data}")

            full_name = data.get('full_name')
            phone = data.get('phone')
            source = data.get('source')
            school = data.get('school')
            answers = data.get('answers', [])

            last_name, first_name, middle_name = (full_name.split() + ["", "", ""])[:3]

            applicant = Applicant.objects.create(
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                phone_number=phone,
                source=source,
                school=school,
                registration_date=datetime.now().date(),
                registration_time=datetime.now().time()
            )

            QuizResult.objects.create(
                answers=answers
            )

            init_excel()
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws1 = wb['Анкеты']
            ws1.append([
                f"{last_name} {first_name} {middle_name}".strip(),
                phone,
                school,
                source,
                
                applicant.application_number,
                str(applicant.registration_date),
                str(applicant.registration_time)[:8]
            ])

            ws2 = wb['Викторины']
            ws2.append([
                f"{last_name} {first_name} {middle_name}".strip(),
                json.dumps(answers, ensure_ascii=False),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ])

            wb.save(EXCEL_PATH)
            print(f"===> Данные записаны в Excel: {EXCEL_PATH}")

            return JsonResponse({'status': 'success'})
        except Exception as e:
            print(f"===> Ошибка: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@csrf_exempt
def save_quiz_result(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            answers = data.get('answers', [])
            result = QuizResult.objects.create(answers=answers)
            return JsonResponse({'status': 'success', 'id': result.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


def quiz_page(request):
    return render(request, 'statistics/quiz.html')


def submit_form(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        phone = request.POST.get('phone')
        source = request.POST.get('source')
        school = request.POST.get('school')
        city = request.POST.get('city')
        
        last_name, first_name, middle_name = (full_name.split() + ["", "", ""])[:3]

        Applicant.objects.create(
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
            phone_number=phone,
            school=school,
            source=source,
            city=city,
            application_number=generate_unique_id(),
            registration_date=datetime.now().date(),
            registration_time=datetime.now().time()
        )

        return render(request, 'statistics/form.html', {'success': True})

    return render(request, 'statistics/form.html')


def generate_unique_id():
    last_applicant = Applicant.objects.order_by('-id').first()
    next_id = 1 if not last_applicant else last_applicant.id + 1
    return f"{next_id:03}"


@login_required
def statistic_view(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("У вас нет доступа к этой странице.")
    
    applicants = Applicant.objects.all()
    return render(request, 'statistics/statistic.html', {'applicants': applicants})
